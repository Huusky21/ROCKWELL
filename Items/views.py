from datetime import date, datetime

from django.conf import settings
from django.contrib import messages
from django.core.mail import send_mail
from django.db import models, transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

import openpyxl
from openpyxl.utils import get_column_letter

from .forms import CapturarForm, RegistroForm
from .models import (
    Accesorio,
    ConteoHuacal,
    DestinatarioAlertaHuacal,
    Entrada,
    Huacal,
    MensajeAlertaHuacal,
    Registro,
    SOItem,
)


REGISTROS_POR_PAGINA = 40


def _filtrar_registros(request):
    q = (request.GET.get('q') or '').strip()
    desde = _parsear_fecha(request.GET.get('desde'))
    hasta = _parsear_fecha(request.GET.get('hasta'))

    qs = Registro.objects.select_related('so_item').order_by('-creado')
    if q:
        qs = qs.filter(
            models.Q(so_item__so_item__icontains=q)
            | models.Q(so_item__material__icontains=q)
        )
    if desde:
        qs = qs.filter(creado__date__gte=desde)
    if hasta:
        qs = qs.filter(creado__date__lte=hasta)
    return qs, {
        'q': q,
        'desde': desde.isoformat() if desde else '',
        'hasta': hasta.isoformat() if hasta else '',
    }


def registros_listado(request):
    qs, filtros = _filtrar_registros(request)
    total = qs.count()
    registros = list(qs[:REGISTROS_POR_PAGINA])

    so_items_count = SOItem.objects.count()
    cantidad_total = sum(r.cantidad for r in registros)

    return render(
        request,
        'Items/registros_listado.html',
        {
            'registros': registros,
            'total': total,
            'mostrando': len(registros),
            'por_pagina': REGISTROS_POR_PAGINA,
            'so_items_count': so_items_count,
            'cantidad_total': cantidad_total,
            **filtros,
        },
    )


def capturar(request):
    if request.method == 'POST':
        form = CapturarForm(request.POST)
        if form.is_valid():
            so_item = form.cleaned_data['so_item']
            cantidad = form.cleaned_data['cantidad']
            huacal = form.cleaned_data.get('huacal')
            accesorio = form.cleaned_data.get('accesorio')
            usuario = request.user if request.user.is_authenticated else None
            registro = Registro.objects.filter(so_item=so_item).order_by('-creado').first()
            if registro:
                registro.cantidad += cantidad
                if accesorio:
                    registro.accesorio = accesorio
                registro.save()
                Entrada.objects.create(
                    registro=registro,
                    cantidad=cantidad,
                    huacal=huacal,
                    accesorio=accesorio,
                    usuario=usuario,
                )
                messages.success(request, f'Se sumó {cantidad} al registro existente de {so_item}.')
            else:
                nuevo_registro = form.save(commit=False)
                nuevo_registro.usuario = usuario
                if accesorio:
                    nuevo_registro.accesorio = accesorio
                nuevo_registro.save()
                Entrada.objects.create(
                    registro=nuevo_registro,
                    cantidad=cantidad,
                    huacal=huacal,
                    accesorio=accesorio,
                    usuario=usuario,
                )
                messages.success(request, 'Captura guardada correctamente.')
            return redirect('registros_listado')
    else:
        form = CapturarForm()

    return render(
        request,
        'Items/capturar.html',
        {
            'form': form,
            'so_items': SOItem.objects.all(),
            'accesorios': Accesorio.objects.all(),
        },
    )


def export_excel(request):
    qs, _ = _filtrar_registros(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros'

    headers = ['SO Item', 'Cantidad', 'Usuario', 'Fecha y hora']
    ws.append(headers)

    for registro in qs:
        usuario = registro.usuario.username if registro.usuario else 'Anónimo'
        ws.append([
            str(registro.so_item),
            registro.cantidad,
            usuario,
            registro.creado.strftime('%d/%m/%Y %H:%M:%S'),
        ])

    for i, column_width in enumerate([40, 15, 25, 24], start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="registros_consultas.xlsx"'
    wb.save(response)
    return response


def registro_detalle(request, pk):
    registro = get_object_or_404(Registro.objects.select_related('so_item', 'usuario'), pk=pk)
    entradas = registro.entradas.select_related('usuario').order_by('-creado')
    return render(request, 'Items/registro_detalle.html', {'registro': registro, 'entradas': entradas})


def editar_registro(request, pk):
    registro = get_object_or_404(Registro, pk=pk)
    if request.method == 'POST':
        form = RegistroForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro actualizado correctamente.')
            return redirect('registros_listado')
    else:
        form = RegistroForm(instance=registro)
    return render(request, 'Items/registro_editar.html', {'form': form, 'registro': registro})


def registro_eliminar(request, pk):
    registro = get_object_or_404(Registro.objects.select_related('so_item'), pk=pk)
    if request.method == 'POST':
        registro.delete()
        messages.success(request, 'Registro eliminado.')
        return redirect('registros_listado')
    return render(request, 'Items/registro_eliminar.html', {'registro': registro})


def _enviar_alerta_huacales_cero(huacales_en_cero, fecha):
    destinatarios = list(
        DestinatarioAlertaHuacal.objects.filter(activo=True).values_list('email', flat=True)
    )
    if not destinatarios:
        return 0
    mensaje = MensajeAlertaHuacal.obtener()
    listado = '\n'.join(
        f'- {h.pn} ({h.get_tipo_display()}) {h.descripcion}'.rstrip()
        for h in huacales_en_cero
    )
    cuerpo = (
        f'{mensaje.cuerpo}\n\n'
        f'Fecha de captura: {fecha:%d/%m/%Y}\n'
        f'Huacales en 0:\n{listado}\n'
    )
    send_mail(
        subject=mensaje.asunto,
        message=cuerpo,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=destinatarios,
        fail_silently=True,
    )
    return len(destinatarios)


def _parsear_fecha(valor):
    if not valor:
        return None
    try:
        return datetime.strptime(valor, '%Y-%m-%d').date()
    except ValueError:
        return None


def huacales_inventario(request):
    tipo = (request.GET.get('tipo') or '').strip()
    q = (request.GET.get('q') or '').strip()
    estado = (request.GET.get('estado') or 'activos').strip()

    huacales = Huacal.objects.all()
    if estado == 'activos':
        huacales = huacales.filter(activo=True)
    elif estado == 'inactivos':
        huacales = huacales.filter(activo=False)
    if tipo in (Huacal.PALLET, Huacal.CRATE):
        huacales = huacales.filter(tipo=tipo)
    if q:
        huacales = huacales.filter(
            models.Q(pn__icontains=q) | models.Q(descripcion__icontains=q)
        )
    huacales = list(huacales.order_by('tipo', 'pn'))

    ultimos = {}
    if huacales:
        conteos = (
            ConteoHuacal.objects
            .filter(huacal__in=huacales)
            .order_by('huacal_id', '-fecha')
        )
        seen = set()
        for c in conteos:
            if c.huacal_id in seen:
                continue
            seen.add(c.huacal_id)
            ultimos[c.huacal_id] = c

    filas = [{'huacal': h, 'ultimo': ultimos.get(h.id)} for h in huacales]

    total_activos = Huacal.objects.filter(activo=True).count()
    total_pallets = Huacal.objects.filter(activo=True, tipo=Huacal.PALLET).count()
    total_crates = Huacal.objects.filter(activo=True, tipo=Huacal.CRATE).count()
    fechas_capturadas = ConteoHuacal.objects.values('fecha').distinct().count()
    en_cero = sum(1 for u in ultimos.values() if u and u.cantidad == 0)

    return render(
        request,
        'Items/huacales_inventario.html',
        {
            'filas': filas,
            'q': q,
            'tipo': tipo,
            'estado': estado,
            'tipo_choices': Huacal.TIPO_CHOICES,
            'stats': {
                'total': total_activos,
                'pallets': total_pallets,
                'crates': total_crates,
                'fechas': fechas_capturadas,
                'en_cero': en_cero,
            },
        },
    )


def huacal_detalle(request, pk):
    huacal = get_object_or_404(Huacal, pk=pk)

    if request.method == 'POST':
        accion = request.POST.get('accion')
        if accion == 'capturar':
            fecha = _parsear_fecha(request.POST.get('fecha'))
            raw_cantidad = request.POST.get('cantidad', '')
            try:
                cantidad = int(raw_cantidad)
            except (TypeError, ValueError):
                cantidad = -1
            if not fecha:
                messages.error(request, 'Fecha inválida.')
            elif cantidad < 0:
                messages.error(request, 'Cantidad inválida.')
            else:
                usuario = request.user if request.user.is_authenticated else None
                ConteoHuacal.objects.update_or_create(
                    huacal=huacal,
                    fecha=fecha,
                    defaults={'cantidad': cantidad, 'usuario': usuario},
                )
                messages.success(
                    request,
                    f'Conteo guardado: {cantidad} en {fecha:%d/%m/%Y}.',
                )
                if cantidad == 0:
                    enviados = _enviar_alerta_huacales_cero([huacal], fecha)
                    if enviados:
                        messages.warning(
                            request,
                            f'Se notificó a {enviados} destinatario(s).',
                        )
                    else:
                        messages.warning(
                            request,
                            'Cantidad en 0, pero no hay destinatarios configurados.',
                        )
        elif accion == 'eliminar_conteo':
            try:
                conteo_pk = int(request.POST.get('conteo_pk', ''))
            except (TypeError, ValueError):
                conteo_pk = None
            if conteo_pk is not None:
                ConteoHuacal.objects.filter(pk=conteo_pk, huacal=huacal).delete()
                messages.success(request, 'Conteo eliminado.')
        return redirect('huacal_detalle', pk=huacal.pk)

    conteos = list(huacal.conteos.select_related('usuario').order_by('-fecha'))
    ultimo = conteos[0] if conteos else None
    return render(
        request,
        'Items/huacal_detalle.html',
        {
            'huacal': huacal,
            'conteos': conteos,
            'ultimo': ultimo,
            'hoy': date.today().isoformat(),
        },
    )


def _guardar_huacal_desde_post(request, instancia=None):
    pn = (request.POST.get('pn') or '').strip()
    if not pn:
        messages.error(request, 'El PN es obligatorio.')
        return False
    if Huacal.objects.exclude(pk=instancia.pk if instancia else None).filter(pn=pn).exists():
        messages.error(request, f'Ya existe un huacal con PN {pn}.')
        return False
    tipo = request.POST.get('tipo') or Huacal.PALLET
    if tipo not in (Huacal.PALLET, Huacal.CRATE):
        tipo = Huacal.PALLET
    try:
        consumo = int(request.POST.get('consumo_semanal') or 0)
    except (TypeError, ValueError):
        consumo = 0
    activo = request.POST.get('activo') == 'on'

    if instancia is None:
        instancia = Huacal(pn=pn)
    instancia.pn = pn
    instancia.descripcion = (request.POST.get('descripcion') or '').strip()
    instancia.medidas = (request.POST.get('medidas') or '').strip()
    instancia.consumo_semanal = max(consumo, 0)
    instancia.tipo = tipo
    instancia.activo = activo
    instancia.save()
    return instancia


def huacal_crear(request):
    if request.method == 'POST':
        if 'activo' not in request.POST:
            request.POST = request.POST.copy()
            request.POST['activo'] = 'on'
        nuevo = _guardar_huacal_desde_post(request)
        if nuevo:
            messages.success(request, f'Huacal {nuevo.pn} creado.')
            return redirect('huacales_inventario')
    return render(
        request,
        'Items/huacal_form.html',
        {
            'huacal': None,
            'tipo_choices': Huacal.TIPO_CHOICES,
            'modo': 'crear',
        },
    )


def huacal_editar(request, pk):
    huacal = get_object_or_404(Huacal, pk=pk)
    if request.method == 'POST':
        resultado = _guardar_huacal_desde_post(request, instancia=huacal)
        if resultado:
            messages.success(request, f'Huacal {huacal.pn} actualizado.')
            return redirect('huacales_inventario')
    return render(
        request,
        'Items/huacal_form.html',
        {
            'huacal': huacal,
            'tipo_choices': Huacal.TIPO_CHOICES,
            'modo': 'editar',
        },
    )


def huacal_eliminar(request, pk):
    huacal = get_object_or_404(Huacal, pk=pk)
    if request.method == 'POST':
        pn = huacal.pn
        huacal.delete()
        messages.success(request, f'Huacal {pn} eliminado.')
        return redirect('huacales_inventario')
    return render(request, 'Items/huacal_eliminar.html', {'huacal': huacal})


def huacales_alertas(request):
    mensaje = MensajeAlertaHuacal.obtener()

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'agregar_destinatario':
            email = (request.POST.get('email') or '').strip()
            nombre = (request.POST.get('nombre') or '').strip()
            if not email:
                messages.error(request, 'El correo es obligatorio.')
            else:
                _, creado = DestinatarioAlertaHuacal.objects.get_or_create(
                    email=email,
                    defaults={'nombre': nombre, 'activo': True},
                )
                if creado:
                    messages.success(request, f'Destinatario {email} agregado.')
                else:
                    messages.info(request, f'{email} ya existía.')

        elif accion == 'eliminar_destinatario':
            try:
                pk = int(request.POST.get('pk', ''))
            except (TypeError, ValueError):
                pk = None
            if pk is not None:
                DestinatarioAlertaHuacal.objects.filter(pk=pk).delete()
                messages.success(request, 'Destinatario eliminado.')

        elif accion == 'guardar_mensaje':
            mensaje.asunto = (request.POST.get('asunto') or '').strip() or mensaje.asunto
            mensaje.cuerpo = request.POST.get('cuerpo') or mensaje.cuerpo
            mensaje.save()
            messages.success(request, 'Mensaje de alerta actualizado.')

        return redirect('huacales_alertas')

    destinatarios = DestinatarioAlertaHuacal.objects.all()
    return render(
        request,
        'Items/huacales_alertas.html',
        {'mensaje': mensaje, 'destinatarios': destinatarios},
    )
